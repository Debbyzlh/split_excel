import streamlit as st
import os
import zipfile
import io
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

# Style-preserving cell copy
def copy_cell(source_cell, target_cell):
    # Preserve formulas if they exist
    if source_cell.data_type == "f":
        target_cell.value = f"={source_cell.value}"  # must explicitly prefix with '='
    else:
        target_cell.value = source_cell.value

    if source_cell.has_style:
        if source_cell.font:
            target_cell.font = source_cell.font.copy()
        if source_cell.border:
            target_cell.border = source_cell.border.copy()
        if source_cell.fill:
            target_cell.fill = source_cell.fill.copy()
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.protection:
            target_cell.protection = source_cell.protection.copy()
        if source_cell.alignment:
            target_cell.alignment = source_cell.alignment.copy()

# Copy a row block from one worksheet to another
def copy_rows(src_ws, tgt_ws, row_start, row_end, tgt_start):
    for i, row in enumerate(src_ws.iter_rows(min_row=row_start, max_row=row_end), start=tgt_start):
        for j, cell in enumerate(row, start=1):
            copy_cell(cell, tgt_ws.cell(row=i, column=j))

st.title("📊 Excel Sheet Splitter (for Grace)")

uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])
if uploaded_file:
    wb = load_workbook(uploaded_file)
    sheet_name = st.selectbox("Select a sheet to split", wb.sheetnames)
    ws = wb[sheet_name]

    st.write("### 共享表头区域 (0行开始, 包含逻辑)")
    header_start = st.number_input("起始行", min_value=0, value=0)
    header_end = st.number_input("结束行 (包含)", min_value=header_start, value=1)

    st.subheader("数据预览")
    preview_data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=1, max_row=10)]
    st.dataframe(preview_data)

    # Extract header row
    sample_header = [cell.value for cell in ws[header_end + 1]]
    col_options = [f"{idx}: {val}" for idx, val in enumerate(sample_header)]

    split_col = st.selectbox("按列拆分 (index: value)", col_options)
    name_col = st.selectbox("文件命名列 (index: value)", col_options)

    split_col_index = int(split_col.split(":")[0])
    name_col_index = int(name_col.split(":")[0])

    if st.button("拆分并下载"):
        output_buffer = io.BytesIO()
        with zipfile.ZipFile(output_buffer, "w") as zipf:
            data_start = header_end + 2
            rows = list(ws.iter_rows(min_row=data_start))

            groups = {}
            for row in rows:
                key = row[split_col_index].value
                if key not in groups:
                    groups[key] = []
                groups[key].append(row)

            for key, group_rows in groups.items():
                name = str(group_rows[0][name_col_index].value).strip() if group_rows[0][name_col_index].value else "Unnamed"
                original_sheet_name = ws.title.strip() if ws.title else "Sheet"
                safe_name = f"{original_sheet_name}_{name}".replace("/", "_").replace("\\", "_")[:80]

                new_wb = Workbook()
                default_ws = new_wb.active
                new_wb.remove(default_ws)
                new_ws = new_wb.create_sheet(title=original_sheet_name[:31])

                if ws.freeze_panes:
                    new_ws.freeze_panes = ws.freeze_panes

                copy_rows(ws, new_ws, header_start + 1, header_end + 1, 1)

                start_row = header_end - header_start + 2
                for i, row in enumerate(group_rows, start=start_row):
                    for j, cell in enumerate(row, start=1):
                        copy_cell(cell, new_ws.cell(row=i, column=j))

                file_buf = io.BytesIO()
                new_wb.save(file_buf)
                zipf.writestr(f"{safe_name}.xlsx", file_buf.getvalue())

        st.download_button(
            label="📥 下载所有拆分文件 (.zip)",
            data=output_buffer.getvalue(),
            file_name="split_excel_files.zip",
            mime="application/zip"
        )